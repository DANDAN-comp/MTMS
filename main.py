import re
from flask import Flask, render_template, request, jsonify, redirect, flash, url_for, send_file
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from io import BytesIO
import openpyxl
import os
import pandas as pd
from flask_caching import Cache
from datetime import datetime, timedelta
import pytz
import sqlite3
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column
from sqlalchemy import text, Integer, String, Float, or_
import time
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from flask import jsonify
from werkzeug.utils import secure_filename




app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)


app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///MouldTool.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False


# Configure the additional database
app.config['SQLALCHEMY_BINDS'] = {
    'new_db': 'sqlite:///NewToolLog.db',
    'jig_db': 'sqlite:///JigLog.db',
    'analysis_db': 'sqlite:///AnalysisLog.db',

}


db = SQLAlchemy(app)


class Part(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    part_number = db.Column(db.String(100), unique=True, nullable=False)
    tool_location = db.Column(db.String(250))
    jig_location = db.Column(db.String(250))
    plug_location = db.Column(db.String(250))
    fit_check_location = db.Column(db.String(250))

with app.app_context():
    db.create_all()



class ShopFloorToolLog(db.Model):
    __bind_key__ = 'new_db'
    id = db.Column(db.Integer, primary_key=True)
    part_number = db.Column(db.String(100), unique=True, nullable=False)
    status = db.Column(db.String(100))
    timestamp = db.Column(db.DateTime, nullable=False)


with app.app_context():
    db.create_all()


class JigLog(db.Model):
    __bind_key__ = 'jig_db'
    id = db.Column(db.Integer, primary_key=True)
    jig_number = db.Column(db.String(100), unique=True, nullable=False)
    status = db.Column(db.String(100))
    timestamp = db.Column(db.DateTime, nullable=False)


with app.app_context():
    db.create_all()

class DataAnalysisLogMould(db.Model):
    __bind_key__ = 'analysis_db'
    __tablename__ = 'data_analysis_log_mould'
    id = db.Column(db.Integer, primary_key=True)
    part_number = db.Column(db.String(100), nullable=False)
    time_sent_to_shop_floor = db.Column(db.DateTime, nullable=False)
    time_retrieved_back_to_storage = db.Column(db.DateTime, nullable=False)
    total_time_on_shop_floor = db.Column(db.Interval, nullable=False)

class DataAnalysisLogJig(db.Model):
    __bind_key__ = 'analysis_db'
    __tablename__ = 'data_analysis_log_jig'
    id = db.Column(db.Integer, primary_key=True)
    jig_number = db.Column(db.String(100), nullable=False)
    time_sent_to_shop_floor = db.Column(db.DateTime, nullable=False)
    time_retrieved_back_to_storage = db.Column(db.DateTime, nullable=False)
    total_time_on_shop_floor = db.Column(db.Interval, nullable=False)

with app.app_context():
    db.create_all()
@app.route('/save_tool', methods=['GET', 'POST'])
def save_tool():
    if request.method == 'POST':
        form_type = request.form.get('form_type')  # either 'mould' or 'jig'
        status = request.form.get('statuss')
        uk_timezone = pytz.timezone('Europe/London')
        timestamp_dt = datetime.now(uk_timezone).replace(tzinfo=None)  # Convert to naive datetime

        if not status:
            flash('Please select a status before saving.', 'warning')
            return redirect(url_for('save_tool'))

        if form_type == 'mould':
            raw = request.form.get('part_numberr', '')
            parts = [p.strip() for p in raw.split(',') if p.strip()]
            if not parts:
                flash('Please scan at least one mould tool number.', 'warning')
                return redirect(url_for('save_tool'))
            for p in parts:
                existing_part = ShopFloorToolLog.query.filter_by(part_number=p).first()
                if status == 'Retrieved Back to Storage' and not existing_part:
                    flash(f'Mould tool number {p} is not present in shop floor to be retrieved.', 'warning')
                    continue
                if existing_part:
                    if status == 'Retrieved Back to Storage':
                        time_sent_to_shop_floor = existing_part.timestamp.replace(tzinfo=None)
                        total_time_on_shop_floor = timestamp_dt - time_sent_to_shop_floor
                        new_analysis_log = DataAnalysisLogMould(
                            part_number=p,
                            time_sent_to_shop_floor=time_sent_to_shop_floor,
                            time_retrieved_back_to_storage=timestamp_dt,
                            total_time_on_shop_floor=total_time_on_shop_floor
                        )
                        db.session.add(new_analysis_log)
                        db.session.delete(existing_part)
                        flash(f'Mould tool number {p} retrieved back to storage and deleted.', 'success')
                    else:
                        flash(f'Mould tool number already exists.', 'warning')
                    continue
                new_log = ShopFloorToolLog(
                    part_number=p,
                    status=status,
                    timestamp=timestamp_dt
                )
                db.session.add(new_log)
                flash(f'New tool saved', 'success')

        elif form_type == 'jig':
            raw = request.form.get('jig_number', '')
            jigs = [j.strip() for j in raw.split(',') if j.strip()]
            if not jigs:
                flash('Please scan at least one jig number.', 'warning')
                return redirect(url_for('save_tool'))
            for j in jigs:
                existing_jig = JigLog.query.filter_by(jig_number=j).first()
                if status == 'Retrieved Back to Storage' and not existing_jig:
                    flash(f'Jig number {j} is not present in shop floor to be retrieved.', 'warning')
                    continue
                if existing_jig:
                    if status == 'Retrieved Back to Storage':
                        time_sent_to_shop_floor = existing_jig.timestamp.replace(tzinfo=None)
                        total_time_on_shop_floor = timestamp_dt - time_sent_to_shop_floor
                        new_analysis_log = DataAnalysisLogJig(
                            jig_number=j,
                            time_sent_to_shop_floor=time_sent_to_shop_floor,
                            time_retrieved_back_to_storage=timestamp_dt,
                            total_time_on_shop_floor=total_time_on_shop_floor
                        )
                        db.session.add(new_analysis_log)
                        db.session.delete(existing_jig)
                        flash(f'Jig number {j} retrieved back to storage and deleted.', 'success')
                    else:
                        flash(f'Jig number already exists.', 'warning')
                    continue
                new_log = JigLog(
                    jig_number=j,
                    status=status,
                    timestamp=timestamp_dt
                )
                db.session.add(new_log)
                flash(f'New jig saved', 'success')

        else:
            flash('Unknown form submission.', 'error')
            return redirect(url_for('save_tool'))

        db.session.commit()
        return redirect(url_for('save_tool'))

    mould_logs = ShopFloorToolLog.query.order_by(ShopFloorToolLog.timestamp.desc()).all()
    jig_logs = JigLog.query.order_by(JigLog.timestamp.desc()).all()
    return render_template('save_tool.html', mould_logs=mould_logs, jig_logs=jig_logs)

@app.route('/update_tools', methods=['POST'])
def update_tools():
    tools = request.form.to_dict(flat=False)

    # Loop over all rows using the index keys
    row_count = len(tools['tools[0][part_number]'])  # assuming consistent length

    for i in range(row_count):
        part_number = tools[f'tools[{i}][part_number]'][0]
        tool_location = tools[f'tools[{i}][tool_location]'][0]
        jig_location = tools[f'tools[{i}][jig_location]'][0]
        plug_location = tools[f'tools[{i}][plug_location]'][0]
        fit_check_location = tools[f'tools[{i}][fit_check_location]'][0]

        part = Part.query.filter_by(part_number=part_number).first()
        if part:
            part.tool_location = tool_location
            part.jig_location = jig_location
            part.plug_location = plug_location
            part.fit_check_location = fit_check_location
        else:
            flash(f"Part number {part_number} not found.", "error")

    db.session.commit()
    flash("Tool locations updated successfully!", "success")
    return redirect(url_for('view_tools'))  # Change 'view_tools' if your endpoint differs

def format_timedelta_days_hours_minutes(td):
    total_seconds = int(td.total_seconds())
    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, seconds = divmod(remainder, 60)

    formatted_time = ""
    if days > 0:
        formatted_time += f"{days} day{'s' if days > 1 else ''} "
    if hours > 0:
        formatted_time += f"{hours} hr{'s' if hours > 1 else ''} "
    if minutes > 0:
        formatted_time += f"{minutes} min{'s' if minutes > 1 else ''}"

    return formatted_time.strip()

@app.route('/import_excel', methods=['POST'])
def import_excel():
    file = request.files.get('file')

    if not file:
        return "No file uploaded", 400

    filename = secure_filename(file.filename)
    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        return "Unsupported file format", 400

    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)

        expected_columns = ['PART NUMBER', 'TOOL LOCATION', 'JIG LOCATION', 'PLUG LOCATION', 'FIT CHECK LOCATION']
        if not all(col in df.columns for col in expected_columns):
            return "Missing expected columns in the file", 400

        # Drop rows with empty or NaN 'PART NUMBER'
        df = df.dropna(subset=['PART NUMBER'])

        # Strip whitespace and ensure string type
        df['PART NUMBER'] = df['PART NUMBER'].astype(str).str.strip()

        for _, row in df.iterrows():
            part_number = row['PART NUMBER']
            if not part_number:
                continue  # Skip empty part numbers

            part = Part.query.filter_by(part_number=part_number).first()
            if part:
                part.tool_location = row['TOOL LOCATION']
                part.jig_location = row['JIG LOCATION']
                part.plug_location = row['PLUG LOCATION']
                part.fit_check_location = row['FIT CHECK LOCATION']
            else:
                new_part = Part(
                    part_number=part_number,
                    tool_location=row['TOOL LOCATION'],
                    jig_location=row['JIG LOCATION'],
                    plug_location=row['PLUG LOCATION'],
                    fit_check_location=row['FIT CHECK LOCATION']
                )
                db.session.add(new_part)

        db.session.commit()
        return "Import successful"

    except Exception as e:
        print("Import error:", e)
        return f"Failed to import file: {e}", 500
@app.route('/export_csv', methods=['GET'])
def export_csv():
    # Query all parts from the database
    parts = Part.query.all()

    # Prepare data for Excel
    data = [{
        'PART NUMBER': part.part_number,
        'TOOL LOCATION': part.tool_location,
        'JIG LOCATION': part.jig_location,
        'PLUG LOCATION': part.plug_location,
        'FIT CHECK LOCATION': part.fit_check_location
    } for part in parts]

    # Create DataFrame and Excel file in memory
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ToolData')

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='ToolData.xlsx',
        as_attachment=True
    )
@app.route('/data_analysis')
def data_analysis():
    mould_logs = DataAnalysisLogMould.query.order_by(DataAnalysisLogMould.time_retrieved_back_to_storage.desc()).all()
    jig_logs = DataAnalysisLogJig.query.order_by(DataAnalysisLogJig.time_retrieved_back_to_storage.desc()).all()
    return render_template('data_analysis.html', mould_logs=mould_logs, jig_logs=jig_logs)

@app.route('/clear_data_analysis', methods=['POST'])
def clear_data_analysis():
    try:
        DataAnalysisLogMould.query.delete()
        DataAnalysisLogJig.query.delete()
        db.session.commit()
        flash('All data analysis logs have been cleared.', 'success')
    except OperationalError:
        db.session.rollback()
        flash('Database is locked. Please try again later.', 'danger')
    return redirect(url_for('data_analysis'))


@app.route('/split1test')
def split1test():
    moulds_in_use = ShopFloorToolLog.query.filter_by(status='Sent to Shop Floor').count()
    jigs_in_use = JigLog.query.filter_by(status='Sent to Shop Floor').count()

    print("Moulds in use:", moulds_in_use)
    print("Jigs in use:", jigs_in_use)

    return render_template('SPLIT 1 TEST.html',
                           moulds_in_use=5,
                           jigs_in_use=3)



@app.route('/api/jig_logs')
def get_jig_logs():
    jigs = JigLog.query.all()
    jig_data = [{
        'id': jig.id,
        'jig_number': jig.jig_number,
        'status': jig.status,
        'timestamp': jig.timestamp.isoformat()
    } for jig in jigs]
    return jsonify(jig_data)

@app.template_filter('format_duration')
def format_duration(td):
    total_seconds = int(td.total_seconds())
    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    minutes = (total_seconds % 3600) // 60

    parts = []
    if days:
        parts.append(f"{days} day{'s' if days > 1 else ''}")
    if hours:
        parts.append(f"{hours} hr{'s' if hours > 1 else ''}")
    if minutes:
        parts.append(f"{minutes} minute{'s' if minutes > 1 else ''}")
    return ' '.join(parts) if parts else '0 minutes'


@app.route('/update_tool_status/<int:log_id>', methods=['POST'])
def update_tool_status(log_id):
    data = request.get_json()
    status = data.get('status')
    log = ShopFloorToolLog.query.get(log_id)

    if log:
        log.status = status
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False})


@app.route('/delete_tool_log/<int:log_id>', methods=['POST'])
def delete_tool_log(log_id):
    log = ShopFloorToolLog.query.get(log_id)

    if log:
        db.session.delete(log)
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False})



@app.route('/save-tool', methods=['POST'])
def savetool():
    part_number = request.form['part_number']
    print("Saving part:", part_number)
    tool_location = request.form.get('tool_location')
    jig_location = request.form.get('jig_location')
    plug_location = request.form.get('plug_location')
    fit_check_location = request.form.get('fit_check_location')

    # Check if part already exists (optional)
    existing = Part.query.filter_by(part_number=part_number).first()
    if existing:
        flash('Part already exists!', 'error')
        return redirect(url_for('view_tools'))

    new_part = Part(
        part_number=part_number,
        tool_location=tool_location,
        jig_location=jig_location,
        plug_location=plug_location,
        fit_check_location=fit_check_location
    )
    db.session.add(new_part)
    db.session.commit()
    # Flash success message
    flash('Tool saved successfully!', 'success')

    # Redirect to the index or view tools page
    return redirect(url_for('view_tools'))

@app.route('/update_tool/<int:tool_id>', methods=['POST'])
def update_tool(tool_id):
    data = request.get_json()  # Get the JSON data from the request

    # Debugging - log the received data
    print(f"Received data for tool {tool_id}: {data}")

    tool = Part.query.get(tool_id)
    if not tool:
        return jsonify({'error': 'Tool not found'}), 404

    # Update the tool attributes with the received data
    tool.tool_location = data.get('tool_location', tool.tool_location)
    tool.jig_location = data.get('jig_location', tool.jig_location)
    tool.plug_location = data.get('plug_location', tool.plug_location)
    tool.fit_check_location = data.get('fit_check_location', tool.fit_check_location)

    db.session.commit()

    # Return the updated tool as a response
    return jsonify({
        'id': tool.id,
        'tool_location': tool.tool_location,
        'jig_location': tool.jig_location,
        'plug_location': tool.plug_location,
        'fit_check_location': tool.fit_check_location
    })


@app.route('/delete_tool/<int:tool_id>', methods=['POST'])
def delete_tool(tool_id):
    tool = Part.query.get(tool_id)
    if not tool:
        return jsonify({'error': 'Tool not found'}), 404

    db.session.delete(tool)
    db.session.commit()

    return jsonify({'success': True})
@app.route('/search_tools', methods=['POST'])
def search_tools():
    try:
        part_number = request.form.get('part_number', '').strip()
        tool_location = request.form.get('tool_location', '').strip()
        jig_location = request.form.get('jig_location', '').strip()
        plug_location = request.form.get('plug_location', '').strip()
        fit_check_location = request.form.get('fit_check_location', '').strip()

        filters = []
        if part_number:
            filters.append(Part.part_number.ilike(f"%{part_number}%"))
        if tool_location:
            filters.append(Part.tool_location.ilike(f"%{tool_location}%"))
        if jig_location:
            filters.append(Part.jig_location.ilike(f"%{jig_location}%"))
        if plug_location:
            filters.append(Part.plug_location.ilike(f"%{plug_location}%"))
        if fit_check_location:
            filters.append(Part.fit_check_location.ilike(f"%{fit_check_location}%"))

        if not filters:
            return jsonify({'tools': []})

        tools = Part.query.filter(or_(*filters)).all()

        return jsonify({
            'tools': [
                {
                    'id': tool.id,
                    'part_number': tool.part_number,
                    'tool_location': tool.tool_location,
                    'jig_location': tool.jig_location,
                    'plug_location': tool.plug_location,
                    'fit_check_location': tool.fit_check_location,
                } for tool in tools
            ]
        })

    except SQLAlchemyError as e:
        print("Database error:", e)
        return jsonify({'error': 'Database query failed'}), 500

    except Exception as e:
        print("Unexpected error:", e)
        return jsonify({'error': 'Unexpected server error'}), 500

@app.route('/clear_db', methods=['POST'])
def clear_db():
    # Verify the password securely (you can improve this logic in a production environment)
    correct_password = 'Infy@123'
    data = request.get_json()
    if not data or 'password' not in data:
        return jsonify({'error': 'No password provided'}), 400

    password = data.get('password')

    if password == correct_password:
        try:
            # Clear all data in the Part table
            db.session.query(Part).delete()
            db.session.commit()
            return jsonify({'message': 'Database cleared successfully.'}), 200
        except Exception as e:
            db.session.rollback()  # In case something goes wrong
            return jsonify({'error': 'An error occurred while clearing the database.'}), 500
    else:
        return jsonify({'error': 'Incorrect password. Database not cleared.'}), 403

# Configure cache (using SimpleCache for in-memory caching)
app.config['CACHE_TYPE'] = 'SimpleCache'
cache = Cache(app)


# SharePoint authentication details
site_url = "https://donite1.sharepoint.com/sites/Donite"
username = "daniel@donite.com"
password = "And096521"
file_url_section1 = "/sites/Donite/Shared Documents/Quality/01-QMS/Records/DONITE Production Approvals/PPAR/PPAR.xlsx"  # Adjust with your file URL
file_url_despatch = "/sites/Donite/Shared Documents/Quality/01-QMS/Records/DONITE Production Approvals/PPAR/TOM DASHBOARD.xlsx"


# Function to authenticate and download file from SharePoint
def get_sharepoint_file(file_url):
    try:
        print("Connecting to SharePoint...")
        ctx = ClientContext(site_url).with_credentials(UserCredential(username, password))
        file_stream = BytesIO()
        file = ctx.web.get_file_by_server_relative_url(file_url)
        file.download(file_stream).execute_query()
        file_stream.seek(0)
        return file_stream
    except Exception as e:
        print(f"Error fetching file: {e}")
        raise


# Function to upload the file to SharePoint
def upload_to_sharepoint(file_url, file_content):
    try:
        ctx = ClientContext(site_url).with_credentials(UserCredential(username, password))
        folder_url = "/sites/Donite/Shared Documents/Quality/01-QMS/Records/DONITE Production Approvals/PPAR/"
        folder = ctx.web.get_folder_by_server_relative_url(folder_url)
        target_file=folder.upload_file(os.path.basename(file_url), file_content)
        ctx.execute_query()
        app.logger.info("File uploaded successfully")
    except Exception as e:
        app.logger.error(f"Error uploading file: {e}")
        raise


@app.errorhandler(500)
def internal_error(error):
    return "500 error: An internal server error occurred.", 500


# Optimized function to fetch and process the despatch data
@app.route('/get_despatch_data', methods=['GET'])
def get_despatch_data():
    selected_date = request.args.get('date')

    # Check if the result is cached
    cached_data = cache.get(f'despatch_data_{selected_date}')
    if cached_data:
        return jsonify({'data': cached_data})

    try:
        # Retrieve the file from SharePoint
        file_stream = get_sharepoint_file(file_url_despatch)

        # Read only necessary columns to improve performance
        despatch_columns = ['Sales.SalesOrderDetails.PartID', 'DespatchNote', 'SalesOrderNumber', 'LineNumber', 'DespatchQuantity', 'DespatchDate', 'Stores.DespatchNotes.CustomerID']
        despatch_df = pd.read_excel(file_stream, sheet_name="Stores DespatchNoteItems", usecols=despatch_columns, engine="openpyxl")

        # Read parts data
        parts_df = pd.read_excel(file_stream, sheet_name="Structure Parts", engine="openpyxl")

        # Convert 'PartID' to a dictionary for faster lookup
        parts_dict = parts_df.set_index('PartID')['PartNumber'].to_dict()

        # Apply map() to get 'Part Number' instead of merge
        despatch_df['Part Number'] = despatch_df['Sales.SalesOrderDetails.PartID'].map(parts_dict).fillna('N/A')

        # Add default values for missing columns
        despatch_df['Customer Code'] = despatch_df.get('Customer Code', 'TOM')

        # Filter by selected date if provided
        if selected_date:
            despatch_df['DespatchDate'] = pd.to_datetime(despatch_df['DespatchDate']).dt.strftime('%Y-%m-%d')
            despatch_df = despatch_df[despatch_df['DespatchDate'] == selected_date]

        # Ensure 'Stores.DespatchNotes.CustomerID' exists and filter by CustomerID = 113
        if 'Stores.DespatchNotes.CustomerID' in despatch_df.columns:
            despatch_df = despatch_df[despatch_df['Stores.DespatchNotes.CustomerID'] == 113]
        else:
            app.logger.warning("'Stores.DespatchNotes.CustomerID' column not found in the data.")

        # Select the necessary columns
        despatch_df = despatch_df[['DespatchNote', 'SalesOrderNumber', 'LineNumber', 'Part Number', 'DespatchQuantity', 'Customer Code', 'DespatchDate']]

        # Convert the result to dictionary for JSON response
        data = despatch_df.to_dict(orient='records')

        # Cache the result for future use (timeout: 5 minutes)
        cache.set(f'despatch_data_{selected_date}', data, timeout=5*60)

        return jsonify({'data': data})

    except Exception as e:
        app.logger.error(f"Error fetching despatch data: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_previous_month_data', methods=['GET'])
def get_previous_month_data():
    try:
        # Retrieve the file from SharePoint
        file_stream = get_sharepoint_file(file_url_despatch)

        # Define the necessary columns and load data
        despatch_columns = [
            'Sales.SalesOrderDetails.PartID', 'DespatchNote', 'SalesOrderNumber',
            'LineNumber', 'DespatchQuantity', 'DespatchDate', 'Stores.DespatchNotes.CustomerID'
        ]
        despatch_df = pd.read_excel(
            file_stream,
            sheet_name="Stores DespatchNoteItems",
            usecols=despatch_columns,
            engine="openpyxl"
        )

        # Filter by CustomerID if the column exists
        if 'Stores.DespatchNotes.CustomerID' in despatch_df.columns:
            despatch_df = despatch_df[despatch_df['Stores.DespatchNotes.CustomerID'] == 113]
        else:
            app.logger.warning("'Stores.DespatchNotes.CustomerID' column not found in the data.")

        # Convert DespatchDate to datetime and drop invalid dates
        despatch_df['DespatchDate'] = pd.to_datetime(despatch_df['DespatchDate'], errors='coerce')
        despatch_df.dropna(subset=['DespatchDate'], inplace=True)

        # Calculate previous month's first and last day
        today = datetime.today()
        first_day_this_month = today.replace(day=1)
        last_day_prev_month = first_day_this_month - timedelta(days=1)
        first_day_prev_month = last_day_prev_month.replace(day=1)

        # Create a full date range DataFrame for the previous month
        full_date_range = pd.date_range(first_day_prev_month, last_day_prev_month)
        full_df = pd.DataFrame({'date_str': full_date_range.strftime('%Y-%m-%d')})

        # Filter the despatch data for dates in the previous month
        mask = (despatch_df['DespatchDate'] >= first_day_prev_month) & (despatch_df['DespatchDate'] <= last_day_prev_month)
        prev_month_df = despatch_df.loc[mask].copy()
        prev_month_df['date_str'] = prev_month_df['DespatchDate'].dt.strftime('%Y-%m-%d')

        # Group by date: Sum the DespatchQuantity and count the SalesOrderNumber (as a proxy for sales orders)
        grouped = prev_month_df.groupby('date_str').agg({
            'DespatchQuantity': 'sum',
            'SalesOrderNumber': 'count'
        }).reset_index()

        # Rename columns to match your JS keys
        grouped.rename(
            columns={'SalesOrderNumber': 'sales_orders', 'DespatchQuantity': 'despatch_quantity'},
            inplace=True
        )

        # Merge the full date range with the grouped data; fill missing days with 0
        merged = pd.merge(full_df, grouped, on='date_str', how='left')
        merged['sales_orders'] = merged['sales_orders'].fillna(0).astype(int)
        merged['despatch_quantity'] = merged['despatch_quantity'].fillna(0)

        # Return data in JSON format
        data = merged.to_dict(orient='records')
        return jsonify({'data': data})
    except Exception as e:
        app.logger.error(f"Error fetching previous month data: {e}")
        return jsonify({'error': str(e)}), 500

def get_price_from_donite_sheet(part_no, qty_shipped, regex_search=False):
    file_url = "/sites/Donite/Shared Documents/Quality/01-QMS/Records/DONITE Production Approvals/PPAR/Donite Thermoforming Price List Feb 2022.xlsx"
    file_stream = get_sharepoint_file(file_url)
    print (qty_shipped)
    print (part_no)


    workbook = openpyxl.load_workbook(file_stream, data_only=True)
    sheet = workbook.active

    header_row = 18
    try:
        qty = float(qty_shipped)
    except ValueError:
        print(f"Invalid quantity: {qty_shipped}")
        return "N/A"

    target_range = determine_target_range(qty)
    if not target_range:
        print(f"No target range found for quantity: {qty}")
        return "N/A"

    col_part_no, col_target = find_columns(sheet, header_row, target_range)
    print(f"Part No: {part_no}, Qty Shipped: {qty_shipped}, Target Range: {target_range}")
    print(f"Column Part No: {col_part_no}, Column Target: {col_target}")

    if col_part_no is None or col_target is None:
        print("Column indices not found")
        return "N/A"

    if regex_search:
        # Use regex to find near matches for the part number
        part_no_pattern = re.compile(part_no, re.IGNORECASE)
    else:
        part_no_pattern = re.compile(re.escape(part_no), re.IGNORECASE)

    for row in range(header_row + 1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=col_part_no).value
        if cell_value and part_no_pattern.match(str(cell_value).strip()):
            price = sheet.cell(row=row, column=col_target).value
            print(f"Found Price: {price} for Part No: {part_no} in Row: {row}")
            return str(price) if price is not None else "N/A"

    print(f"No matching part number found for Part No: {part_no}")
    return "N/A"

def determine_target_range(qty):
    if 1 <= qty <= 4:
        return "1 off"
    elif 5 <= qty <= 9:
        return "5 to 9"
    elif 10 <= qty <= 19:
        return "10 to 19"
    elif 20 <= qty <= 29:
        return "20 to 29"
    elif 30 <= qty <= 49:
        return "30 to 49"
    elif 50 <= qty <= 99:
        return "50 to 99"
    elif 100 <= qty <= 199:
        return "100 to 199"
    elif 200 <= qty <= 299:
        return "200 to 299"
    elif qty >= 300:
        return "300+"
    else:
        return None

def find_columns(sheet, header_row, target_range):
    col_part_no = col_target = None
    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=header_row, column=col).value
        if header_value:
            header_value = str(header_value).strip().lower()
            if header_value == "part number":
                col_part_no = col
            if header_value == target_range.lower():
                col_target = col
    return col_part_no, col_target

def clean_part_no(part_no):
    """
    Removes any trailing alphabetical characters from the part number.
    For example, "VT16-05-049-01-HEX" becomes "VT16-05-049-01".
    """
    return re.sub(r'-[A-Za-z]+$', '', part_no)

def delete_row_from_sharepoint(advice_note):
    file_url = "/sites/Donite/Shared Documents/Quality/01-QMS/Records/DONITE Production Approvals/PPAR/Data_Extractor.xlsx"
    file_stream = get_sharepoint_file(file_url)

    workbook = openpyxl.load_workbook(file_stream)
    sheet = workbook.active

    row_to_delete = None
    for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
        cell_value = sheet.cell(row=row, column=1).value  # Assuming Advice Note is in the first column
        if cell_value == advice_note:
            row_to_delete = row
            break

    if row_to_delete:
        sheet.delete_rows(row_to_delete)
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)
        upload_to_sharepoint(file_url, file_stream)
    else:
        raise Exception(f"Advice Note {advice_note} not found")
@app.route("/")
def home():
    return render_template("SPLIT 1 TEST.html")

@app.route('/get_price_from_donite_sheet', methods=['POST'])
def get_price_from_donite_sheet_route():
    data = request.json
    part_no = data.get('partNo')
    qty_shipped = data.get('qtyShipped')
    regex_search = data.get('regex', False)
    price = get_price_from_donite_sheet(part_no, qty_shipped, regex_search)
    return jsonify({"price": price})

@app.route('/delete_row', methods=['POST'])
def delete_row_route():
    data = request.json
    advice_note = data.get('adviceNote')
    if not advice_note:
        return jsonify({"error": "Advice Note is required"}), 400

    try:
        delete_row_from_sharepoint(advice_note)
        return jsonify({"message": "Row deleted successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/data_extractor')
def data_extractor():
    return render_template('copilot.html')

@app.route('/Thompson Aero')
def Thompson_Aero():
    return render_template('TOM.html')

@app.route('/save_pdf_data', methods=['POST'])
def save_pdf_data():
    try:
        # Get the extracted PDF data from the request (sent as JSON)
        data_list = request.get_json()
        if not data_list:
            return jsonify({"error": "No data provided"}), 400

        # Corrected SharePoint file URL for the main data Excel file
        file_url = "/sites/Donite/Shared Documents/Quality/01-QMS/Records/DONITE Production Approvals/PPAR/Data_Extractor.xlsx"

        # Get the Excel file stream from SharePoint
        new_file_stream = get_sharepoint_file(file_url)
        workbook = openpyxl.load_workbook(new_file_stream)
        sheet = workbook.active

        for data in data_list:
            # Retrieve extracted fields (defaulting to "N/A" if not provided)
            advice_note = data.get("Advice Note", "N/A")
            wo_ref = data.get("WO Ref.", "N/A")
            part_no = data.get("Part No.", "N/A")
            qty_shipped = data.get("Qty Shipped", "N/A")
            expected_receipt = data.get("Expected Receipt", "N/A")
            purchase_ref = data.get("Purchase Ref.", "N/A")
            price_from_advice_note = data.get("Price from Advice Note", "N/A")
            # We'll overwrite the Price from Donite Spreadsheet using our helper function.
            part_issue = data.get("Part Issue", "N/A")
            material = data.get("Material", "N/A")
            qty_sheets_sent = data.get("Qty sheets sent", "N/A")
            description = data.get("Description", "N/A")

            # Clean the part number (remove trailing alphabetical letters)
            cleaned_part_no = clean_part_no(part_no) if part_no != "N/A" else "N/A"

            # Ensure Qty Shipped is numeric before price lookup
            try:
                qty_val = float(qty_shipped)
            except ValueError:
                qty_val = None

            if qty_val is None or cleaned_part_no == "N/A":
                price_from_donite_spreadsheet = "N/A"
            else:
                price_from_donite_spreadsheet = get_price_from_donite_sheet(cleaned_part_no, qty_val)

                # Log for debugging
            app.logger.info(f"Part No.: {cleaned_part_no}, Qty: {qty_val}, Price from Donite: {price_from_donite_spreadsheet}")

            # Create the data dictionary to be added/updated in Excel
            pdf_data = {
                "Advice Note": advice_note,
                "WO Ref.": wo_ref,
                "Part No.": cleaned_part_no,
                "Qty Shipped": qty_shipped,
                "Expected Receipt": expected_receipt,
                "Purchase Ref.": purchase_ref,
                "Price from Advice Note": price_from_advice_note,
                "Price from Donite Spreadsheet": price_from_donite_spreadsheet,
                "Part Issue": part_issue,
                "Material": material,
                "Qty sheets sent": qty_sheets_sent,
                "Description": description
            }
            app.logger.debug("PDF Data to be saved: %s", pdf_data)

            # Use fixed column order when appending
            column_order = [
                "Advice Note", "WO Ref.", "Part No.", "Qty Shipped", "Expected Receipt",
                "Purchase Ref.", "Price from Advice Note", "Price from Donite Spreadsheet",
                "Part Issue", "Material", "Qty sheets sent", "Description"
            ]
            sheet.append([pdf_data.get(key, "N/A") for key in column_order])

        # Save the updated workbook into a BytesIO stream
        updated_file_stream = BytesIO()
        workbook.save(updated_file_stream)
        updated_file_stream.seek(0)

        # Upload the updated Excel file back to SharePoint
        upload_to_sharepoint(file_url, updated_file_stream)

        return jsonify({"message": "PDF data saved successfully"}), 200

    except Exception as e:
        app.logger.error(f"Error while saving PDF data: {e}")
        return jsonify({"error": "Failed to save PDF data"}), 500

@app.route('/get_saved_data', methods=['GET'])
def get_saved_data():
    try:
        # Corrected SharePoint file URL
        file_url = "/sites/Donite/Shared Documents/Quality/01-QMS/Records/DONITE Production Approvals/PPAR/Data_Extractor.xlsx"

        # Get the Excel file stream from SharePoint
        file_stream = get_sharepoint_file(file_url)
        workbook = openpyxl.load_workbook(file_stream)
        sheet = workbook.active

        # Read data from the Excel file
        data_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            data = {
                "Advice Note": row[0] if row[0] is not None else "N/A",
                "WO Ref.": row[1] if row[1] is not None else "N/A",
                "Part No.": row[2] if row[2] is not None else "N/A",
                "Qty Shipped": row[3] if row[3] is not None else "N/A",
                "Expected Receipt": row[4] if row[4] is not None else "N/A",
                "Purchase Ref.": row[5] if row[5] is not None else "N/A",
                "Price from Advice Note": row[6] if row[6] is not None else "N/A",
                "Price from Donite Spreadsheet": row[7] if row[7] is not None else "N/A",
                "Part Issue": row[8] if row[8] is not None else "N/A",
                "Material": row[9] if row[9] is not None else "N/A",
                "Qty sheets sent": row[10] if row[10] is not None else "N/A",
                "Description": row[11] if row[11] is not None else "N/A"
            }
            data_list.append(data)

        return jsonify(data_list), 200

    except Exception as e:
        app.logger.error(f"Error while fetching saved data: {e}")
        return jsonify({"error": "Failed to fetch saved data"}), 500



@app.route("/STOCKCHECK")
def stock_check():
    return render_template("SPLIT 2 TEST.html")

@app.route('/view_tools')
def view_tools():
    tools = Part.query.all()
    return render_template('view_tools.html', tools=tools)


@app.route("/PPAR")
def PPAR():
    return render_template("PPAR.html")

# Start the Flask app and the scheduled task in separate threads
if __name__ == '__main__':
    app.run(debug=True)  #